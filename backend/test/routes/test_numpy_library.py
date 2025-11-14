from fastapi.testclient import TestClient
from routes.numpy_library import router  # replace with your module where router is defined
import os
from openpyxl import load_workbook

# Create FastAPI app for testing
from fastapi import FastAPI
app = FastAPI()
app.include_router(router)

client = TestClient(app)

# -----------------------------
# Sample user input
# -----------------------------
sample_input = {
    "firstName": "John",
    "lastName": "Doe",
    "age": 30,
    "income": 5000.0,
    "monthly_expenses": [1000, 2000, 3000, 4000, 5000],
    "investment_goals": "Retirement",
    "risk_tolerance": "Medium",
    "time_horizon": 10
}

# -----------------------------
# Test endpoint response
# -----------------------------
def test_create_financial_returns_excel():
    response = client.post("/numpy-libraries", json=sample_input)
    
    # Check status code
    assert response.status_code == 200
    
    # Check content type (should be Excel file)
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("content-type", "")
    
    # Save response content to temporary file
    temp_file = "test_expenses.xlsx"
    with open(temp_file, "wb") as f:
        f.write(response.content)
    
    # Load workbook and check sheet names
    wb = load_workbook(temp_file)
    sheets = wb.sheetnames
    assert "Expenses Analysis" in sheets
    assert "Statistics" in sheets
    assert "Graph" in sheets

    # Clean up temporary file
    os.remove(temp_file)

# -----------------------------
# Optional: test statistics values
# -----------------------------
def test_statistics_values():
    expenses = sample_input["monthly_expenses"]
    expected_sum = sum(expenses)
    
    response = client.post("/numpy-libraries", json=sample_input)
    temp_file = "test_expenses.xlsx"
    with open(temp_file, "wb") as f:
        f.write(response.content)
    
    wb = load_workbook(temp_file)
    ws_stats = wb["Statistics"]
    
    # Read sum from Excel
    excel_sum = ws_stats["A2"].value  # Assuming Sum is first column
    assert excel_sum == expected_sum
    
    os.remove(temp_file)
