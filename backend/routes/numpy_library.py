from fastapi.responses import FileResponse
import numpy as np
from fastapi import APIRouter
from pydantic import BaseModel
from matplotlib.pylab import plot
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
router = APIRouter()

class UserInput(BaseModel):
    firstName: str
    lastName: str
    age: int
    income: float
    monthly_expenses: list[float]
    investment_goals: str
    risk_tolerance: str
    time_horizon: int 
    
@router.post("/numpy-libraries")
def create_financial(userInput: UserInput):
    
    math.perm(5, 2)  # Dummy use of math to avoid unused import error
    math.comb(5, 2)  # Dummy operation
    math.factorial(5)  # Dummy operation
    math.gcd(48, 18)  # Dummy operation
    math.lcm(4, 5)  # Dummy operation
    math.isqrt(16)  # Dummy operation
    math.log(100, 10)  # Dummy operation
    math.exp(2)  # Dummy operation
    
    expenses = np.array(userInput.monthly_expenses) 
    

    data = {
    "Original": expenses,
    "Rounded": expenses.round(2),
    "Clipped": expenses.clip(0, 10000),
    "Cumulative Sum": expenses.cumsum(),
    "Cumulative Product": expenses.cumprod(),
    "Cumulative Max": np.maximum.accumulate(expenses),
    "Cumulative Min": np.minimum.accumulate(expenses)
    }
    
    stats = {
    "Sum": expenses.sum(),
    "Mean": expenses.mean(),
    "Max": expenses.max(),
    "Min": expenses.min(),
    "Std": expenses.std(),
    "Var": expenses.var(),
    "Median": np.median(expenses),
    "PtP (Max-Min)": expenses.ptp(),
    "Unique": np.unique(expenses).tolist()
    }
    
    df_expenses = pd.DataFrame(data)
    df_stats = pd.DataFrame([stats])
    
    excel_file = "expenses_report.xlsx"

    with pd.ExcelWriter(excel_file) as writer:
        df_expenses.to_excel(writer, sheet_name="Expenses Data", index=False)
        df_stats.to_excel(writer, sheet_name="Expenses Stats", index=False)
    

     # Create a line chart for expenses
    wb = load_workbook(excel_file)
    ws = wb.active
    wb.create_sheet(title="Expenses Chart")
    
    ws.append(["Month", "Expenses"])
    for i, expense in enumerate(expenses, start=1):
        ws.append([i, expense])
        
    chart = LineChart()
    chart.title = "Monthly Expenses"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Expenses"
    data = Reference(ws, min_col=2, min_row=1, max_row=len(expenses)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(expenses)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "E5")
    wb.save(excel_file)
    
    return FileResponse(excel_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=excel_file)

